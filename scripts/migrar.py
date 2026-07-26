"""
Migração das planilhas -> Supabase.

Fontes (o melhor de cada uma):
  - fontes_planilhas.py       -> motoristas, veiculos, fretes, roteiro,
                                 contas_fixas, contas_pagar, saldos_banco
                                 (lê direto das planilhas originais)
  - POSTO 2026.xlsx           -> abastecimentos
  - FROTA e MANUTENÇOES.xlsx  -> documentos dos veículos + manutencoes

Antes lia de `transportadora-demo/js/data.js`, que era uma cópia estática das
planilhas. Esse arquivo foi apagado em 24/07/2026, quando o sistema passou a
viver só no Supabase — e a migração ficou impossível de repetir. Agora lê das
planilhas de verdade, então dá para rodar de novo do zero, inclusive em um
projeto de teste.

Uso:
  python migrar.py --dry-run     # só analisa e mostra o que faria (não grava nada)
  python migrar.py               # insere no Supabase (pula tabelas que já têm dados)
  python migrar.py --recarregar  # apaga o que existe e insere tudo de novo

ATENÇÃO com --recarregar: ele apaga as tabelas antes de inserir, então também
leva embora tudo o que foi criado pelo sistema depois da migração (checklists,
ocorrências, ordens de manutenção, vales lançados pela tela). Rode
backup_banco.py antes.
"""
import re
import sys
from datetime import date, datetime
from pathlib import Path

import openpyxl

PASTA = Path(__file__).resolve().parent.parent          # ...\PhorteAguiar

# caminhos vêm de fontes_planilhas para não divergirem (pasta fontes/ é a
# fonte de verdade; ver fontes/LEIA-ME.md)
from fontes_planilhas import _fonte, XLSX_POSTO

XLSX_FROTA = _fonte("FROTA e MANUTENÇOES.xlsx")

HOJE = date.today().isoformat()

# placas com grafia dupla nas planilhas -> forma canônica
ALIAS_PLACA = {"NNX3195": "NNX3I95"}


# ----------------------------------------------------------------------------
# helpers
# ----------------------------------------------------------------------------
def placa_norm(p):
    if not p:
        return None
    p = re.sub(r"\s+", "", str(p)).upper()
    return ALIAS_PLACA.get(p, p)


def data_iso(v):
    """datetime/date/str -> 'YYYY-MM-DD' ou None."""
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, date):
        return v.isoformat()
    s = str(v).strip()
    if re.match(r"^\d{4}-\d{2}-\d{2}", s):
        return s[:10]
    if re.match(r"^\d{4}-\d{2}$", s):     # "2028-01" -> primeiro dia do mês
        return s + "-01"
    return None


def num(v, casas=2):
    if v is None or v == "":
        return None
    try:
        return round(float(v), casas)
    except (TypeError, ValueError):
        return None


def txt(v):
    if v is None:
        return None
    s = str(v).strip()
    return s if s and s != "**" else None


# ----------------------------------------------------------------------------
# 1) planilhas originais (antes: data.js)
# ----------------------------------------------------------------------------
from fontes_planilhas import carregar_planilhas


def montar_motoristas(db):
    linhas = []
    for m in db["motoristas"]:
        linhas.append({
            "nome": m["nome"],
            "cpf": txt(m.get("cpf")),
            "rg": txt(m.get("rg")),
            "ativo": True,
        })
    return linhas


def montar_veiculos(db, docs_frota, carretas_info):
    """Cavalos do data.js (km/tacógrafo/motorista) + documentos do FROTA.xlsx.
    Carretas da lista do data.js + ano/marca da aba MANUTENÇÃO CARRETAS.
    Retorna (carretas, cavalos) — carretas primeiro por causa do vínculo carreta_placa."""
    carretas = []
    for placa in db["carretas"]:
        p = placa_norm(placa)
        info = carretas_info.get(p, {})
        carretas.append({
            "placa": p,
            "tipo": "carreta",
            "modelo": info.get("marca"),
            "ano_modelo": info.get("ano"),
            "ativo": True,
        })

    cavalos = []
    for v in db["veiculos"]:
        p = placa_norm(v["placa"])
        docs = docs_frota.get(p, {})
        cavalos.append({
            "placa": p,
            "tipo": "cavalo",
            "modelo": docs.get("modelo"),
            "ano_modelo": docs.get("ano_modelo"),
            "cor": docs.get("cor"),
            "renavam": docs.get("renavam"),
            "chassi": docs.get("chassi"),
            "motor": docs.get("motor"),
            "crv_numero": docs.get("crv_numero"),
            "cod_seg_cla": docs.get("cod_seg_cla"),
            "licenciamento_ano": docs.get("licenciamento_ano"),
            "ipva_status": docs.get("ipva_status"),
            "carreta_placa": placa_norm(v.get("carreta")),
            "antt_empresa": txt(v.get("antt")),
            "antt_numero": txt(v.get("anttNum")),
            "mct_numero": txt(v.get("mct")),
            "mct_status": txt(v.get("mctStatus")),
            "tacografo_venc": data_iso(v.get("tacografo")),
            "tacografo_obs": txt(v.get("tacografoObs")),
            "km_atual": v.get("kmAtual"),
            "km_troca": v.get("kmTroca"),
            "media_kml": num(v.get("media"), 3),
            "media_data": data_iso(v.get("mediaData")),
            "_motorista_nome": v.get("motorista"),   # resolvido depois p/ uuid
            "ativo": True,
        })
    return carretas, cavalos


def montar_fretes(db):
    linhas = []
    for f in db["fretes"]:
        pagto_iso = data_iso(f.get("pagto"))
        obs = txt(f.get("obs"))
        if f.get("pagto") and not pagto_iso:          # ex: "PG 23/01"
            obs = f"{obs + ' · ' if obs else ''}pagto: {f['pagto']}"
        linhas.append({
            "data": data_iso(f["data"]),
            "_motorista_nome": f["motorista"],
            "veiculo_placa": placa_norm(f.get("veiculo")),
            "origem": txt(f.get("origem")),
            "destino": txt(f.get("destino")),
            "transportadora": txt(f.get("transportadora")),
            "valor_frete": num(f.get("frete")) or 0,
            "adiantamento": num(f.get("adiant")) or 0,
            "diaria": num(f.get("diaria")) or 0,
            "saldo": num(f.get("saldo")),
            "comissao": num(f.get("comissao")) or 0,
            "pedagio_valor": num(f.get("pedagio")) or 0,
            "pedagio_forma": txt(f.get("pedagioVia")),
            "pedagio_pago_por": "empresa",
            "ciot": txt(f.get("ciot")),
            "banco": txt(f.get("banco")),
            "pagamento_previsto": pagto_iso,
            "pagamento_realizado": pagto_iso if (pagto_iso and pagto_iso <= HOJE) else None,
            "observacao": obs,
        })
    return linhas


def montar_roteiro(db):
    linhas = []
    for r in db["roteiro"]:
        posicao = r.get("posicao") or ""
        if "→" in posicao:
            origem_uf, destino_uf = [x.strip() for x in posicao.split("→", 1)]
        else:
            origem_uf, destino_uf = posicao.strip() or None, None
        linhas.append({
            "data": db.get("hoje") or HOJE,
            "ordem": r.get("ordem"),
            "_motorista_nome": r["motorista"],
            "origem_uf": origem_uf,
            "destino_uf": destino_uf,
            "status": txt(r.get("status")) or "A CONFIRMAR",
            "observacao": txt(r.get("rota")),
        })
    return linhas


def montar_contas_fixas(db):
    linhas = []
    for c in db["contasFixas"]:
        linhas.append({
            "dia_venc": min(max(int(c["dia"]), 1), 31),
            "descricao": c["descricao"],
            "valor": num(c["valor"]) or 0,
            "recorrencia": txt(c.get("recorrencia")) or "MENSAL",
            "origem": c.get("origem") or "empresa",
            "forma_pagto": txt(c.get("forma")),
            "data_fim": data_iso(c.get("fim")),
            "ativa": bool(c.get("ativa", True)),
            "pendente": bool(c.get("pendente", False)),
            "observacao": txt(c.get("obs")),
        })
    return linhas


def montar_contas_pagar(db):
    linhas = []
    for c in db["financeiro"]["contasPagar"]:
        linhas.append({
            "descricao": c["descricao"],
            "valor": num(c["valor"]) or 0,
            "vencimento": data_iso(c["vencimento"]),
            "forma": txt(c.get("forma")),
            "categoria": txt(c.get("categoria")),
            "grupo": txt(c.get("grupo")),
        })
    return linhas


def montar_saldos(db):
    return [{"banco": s["banco"], "saldo": num(s["saldo"]) or 0}
            for s in db["financeiro"]["saldos"]]


def montar_vales(db):
    linhas = []
    for v in db.get("vales", []):
        linhas.append({
            "_motorista_nome": v["motorista"],
            "data": data_iso(v["data"]),
            "valor": num(v["valor"]) or 0,
            "pago": num(v.get("pago")) or 0,
            "descricao": txt(v.get("descricao")),
        })
    return linhas


# ----------------------------------------------------------------------------
# 2) FROTA e MANUTENÇOES.xlsx
# ----------------------------------------------------------------------------
CAMPOS_DOC = {
    "VEÍCULO": "modelo", "VEICULO": "modelo",
    "RENAVAN": "renavam", "RENAVAM": "renavam",
    "ANO/MOD": "ano_modelo",
    "COR": "cor",
    "NR DO CRV": "crv_numero",
    "CÓD SEG CLA": "cod_seg_cla", "COD SEG CLA": "cod_seg_cla",
    "CHASSI": "chassi",
    "MOTOR": "motor",
    "LICENCIAMENTO": "licenciamento_ano",
    "IPVA": "ipva_status",
}


def ler_frota():
    """Retorna (docs_por_placa, manutencoes, carretas_info)."""
    wb = openpyxl.load_workbook(XLSX_FROTA, read_only=True, data_only=True)
    docs, manutencoes, carretas_info = {}, [], {}

    for aba in wb.sheetnames:
        nome = aba.strip()
        if nome.upper().startswith("MANUTEN"):        # aba MANUTENÇÃO CARRETAS
            ws = wb[aba]
            for row in ws.iter_rows(values_only=True):
                placa = placa_norm(row[1] if len(row) > 1 else None)
                if not placa or placa == "PLACA" or not re.match(r"^[A-Z]{3}\d", placa):
                    continue
                carretas_info[placa] = {
                    "ano": txt(row[2]) if len(row) > 2 else None,
                    "marca": txt(row[3]) if len(row) > 3 else None,
                }
                valor = num(row[5]) if len(row) > 5 else None
                desc = txt(row[4]) if len(row) > 4 else None
                if desc and valor:
                    manutencoes.append({
                        "veiculo_placa": placa,
                        "data": data_iso(row[6] if len(row) > 6 else None) or HOJE,
                        "servico": desc,
                        "valor": valor,
                        "categoria": "carreta",
                    })
            continue
        if nome.lower().startswith("planilha"):
            continue

        placa = placa_norm(nome)
        if not re.match(r"^[A-Z]{3}\d", placa or ""):
            continue
        ws = wb[aba]
        doc, header_idx = {}, None
        linhas = list(ws.iter_rows(values_only=True))

        for i, row in enumerate(linhas):
            rotulo = txt(row[0])
            if rotulo:
                chave = CAMPOS_DOC.get(rotulo.upper())
                if chave and len(row) > 1:
                    doc[chave] = txt(row[1])
                if rotulo.upper().startswith("TAC"):
                    doc["tacografo_venc"] = data_iso(row[1]) if len(row) > 1 else None
            # localizar header da tabela de manutenções
            celulas = [str(c).upper() if c else "" for c in row]
            if any("OCORR" in c for c in celulas) and any("DATA" in c for c in celulas):
                header_idx = {n: j for j, c in enumerate(celulas) for n in
                              ["DATA", "KM", "SERVI", "VALOR", "OFICINA", "OS/NF", "PAGAMENTO"]
                              if n in c}
                continue
            if header_idx:
                col = header_idx
                d = data_iso(row[col["DATA"]]) if "DATA" in col and col["DATA"] < len(row) else None
                serv = txt(row[col["SERVI"]]) if "SERVI" in col and col["SERVI"] < len(row) else None
                if not d or not serv:
                    continue
                manutencoes.append({
                    "veiculo_placa": placa,
                    "data": d,
                    "km": int(row[col["KM"]]) if "KM" in col and isinstance(row[col["KM"]], (int, float)) else None,
                    "servico": serv,
                    "valor": num(row[col["VALOR"]]) if "VALOR" in col else None,
                    "oficina": txt(row[col["OFICINA"]]) if "OFICINA" in col else None,
                    "os_nf": txt(row[col["OS/NF"]]) if "OS/NF" in col else None,
                })
        docs[placa] = doc
    return docs, manutencoes, carretas_info


# ----------------------------------------------------------------------------
# 3) POSTO 2026.xlsx
# ----------------------------------------------------------------------------
MES_ABA = {"JAN": "01", "FEV": "02", "MAR": "03", "ABR": "04", "MAI": "05",
           "JUN": "06", "JUL": "07", "AGO": "08", "SET": "09", "OUT": "10",
           "NOV": "11", "DEZ": "12"}


def ler_posto(mapa_motorista_veiculo):
    wb = openpyxl.load_workbook(XLSX_POSTO, read_only=True, data_only=True)
    abastecimentos, sem_dono, corrigidas = [], set(), []

    for aba in wb.sheetnames:
        ws = wb[aba]
        mes_aba = MES_ABA.get(aba.strip().upper()[:3])
        placa_atual = None
        for row in ws.iter_rows(values_only=True):
            c0 = row[0]
            d = data_iso(c0)
            if d:                                    # linha de abastecimento
                if not placa_atual:
                    continue
                # data digitada com mês errado dentro da aba do mês (ex: 2026-10 na aba JANEIRO)
                if mes_aba and d[:4] == "2026" and d[5:7] != mes_aba:
                    d_corrigida = f"{d[:5]}{mes_aba}{d[7:]}"
                    corrigidas.append(f"{aba}: {d} -> {d_corrigida}")
                    d = d_corrigida
                abastecimentos.append({
                    "veiculo_placa": placa_atual,
                    "data": d,
                    "nota_fiscal": txt(row[1]),
                    "valor": num(row[2]) or 0,
                    "litros": num(row[3], 3),
                    "valor_litro": num(row[4], 4),
                    "km_atual": int(row[5]) if isinstance(row[5], (int, float)) else None,
                    "km_rodado": int(row[6]) if isinstance(row[6], (int, float)) else None,
                    "media_kml": num(row[7], 3),
                    "posto": txt(row[8]),
                    "vencimento": data_iso(row[9]) if len(row) > 9 else None,
                    "valor_vencimento": num(row[10]) if len(row) > 10 else None,
                })
                continue
            titulo = txt(c0)
            if titulo and titulo.upper() not in ("DATA",):   # título de bloco
                t = titulo.upper()
                # placa no próprio título ("FULANO - GAK3400")?
                m = re.search(r"[A-Z]{3}\s?\d[A-Z0-9]\d{2}", t)
                if m:
                    placa_atual = placa_norm(m.group(0))
                    continue
                # senão, procura por nome de motorista conhecido
                achou = None
                for nome_mot, placa in mapa_motorista_veiculo.items():
                    if nome_mot.split()[0] in t:
                        achou = placa
                        break
                if achou:
                    placa_atual = achou
                else:
                    sem_dono.add(titulo)
    if corrigidas:
        print("  Datas com mês errado corrigidas pelo mês da aba (POSTO):")
        for c in corrigidas:
            print(f"    - {c}")
    return abastecimentos, sorted(sem_dono)


# ----------------------------------------------------------------------------
# execução
# ----------------------------------------------------------------------------
def preparar_tudo():
    db = carregar_planilhas()
    docs_frota, manutencoes, carretas_info = ler_frota()
    carretas, cavalos = montar_veiculos(db, docs_frota, carretas_info)
    mapa_mot_veic = {m["nome"]: placa_norm(m.get("veiculo")) for m in db["motoristas"]}
    abastecimentos, blocos_sem_dono = ler_posto(mapa_mot_veic)

    dados = {
        "motoristas": montar_motoristas(db),
        "veiculos": carretas + cavalos,
        "fretes": montar_fretes(db),
        "roteiro": montar_roteiro(db),
        "contas_fixas": montar_contas_fixas(db),
        "contas_pagar": montar_contas_pagar(db),
        "saldos_banco": montar_saldos(db),
        "vales": montar_vales(db),
        "manutencoes": manutencoes,
        "abastecimentos": abastecimentos,
    }
    return dados, blocos_sem_dono


ORDEM_INSERCAO = ["motoristas", "veiculos", "fretes", "roteiro", "contas_fixas",
                  "contas_pagar", "saldos_banco", "vales", "manutencoes", "abastecimentos"]
# ordem inversa p/ apagar sem violar chaves estrangeiras
ORDEM_LIMPEZA = ["abastecimentos", "manutencoes", "vales", "roteiro", "fretes",
                 "contas_pagar", "contas_fixas", "saldos_banco", "veiculos", "motoristas"]


def resumo(dados, blocos_sem_dono):
    print("\n=== RESUMO DO QUE FOI LIDO ===")
    for tabela in ORDEM_INSERCAO:
        print(f"  {tabela:16} {len(dados[tabela]):5} registros")
    tot = sum(f["valor_frete"] for f in dados["fretes"])
    print(f"\n  Fretes: total R$ {tot:,.2f}".replace(",", "X").replace(".", ",").replace("X", "."))
    por_mes = {}
    for f in dados["fretes"]:
        por_mes[f["data"][:7]] = por_mes.get(f["data"][:7], 0) + f["valor_frete"]
    for mes in sorted(por_mes):
        v = f"{por_mes[mes]:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
        print(f"    {mes}: R$ {v}")
    if blocos_sem_dono:
        print("\n  AVISO - blocos do POSTO sem motorista/placa identificados (ignorados):")
        for b in blocos_sem_dono:
            print(f"    - {b}")


def inserir_tudo(dados, recarregar):
    from db import supabase  # só conecta quando vai gravar de verdade

    def contar(tabela):
        r = supabase.table(tabela).select("*", count="exact").limit(1).execute()
        return r.count or 0

    if recarregar:
        print("\nLimpando tabelas (--recarregar)...")
        supabase.table("veiculos").update({"carreta_placa": None}).neq("placa", "").execute()
        for tabela in ORDEM_LIMPEZA:
            coluna = "placa" if tabela == "veiculos" else "id"
            supabase.table(tabela).delete().neq(coluna, "00000000-0000-0000-0000-000000000000"
                                                if coluna == "id" else "").execute()
            print(f"  {tabela}: limpa")

    # 1) motoristas -> mapa nome -> uuid
    if contar("motoristas") == 0:
        supabase.table("motoristas").insert(dados["motoristas"]).execute()
        print(f"motoristas: {len(dados['motoristas'])} inseridos")
    else:
        print("motoristas: já tem dados, pulando (use --recarregar p/ recriar)")
    mapa_mot = {m["nome"]: m["id"]
                for m in supabase.table("motoristas").select("id,nome").execute().data}

    def resolver_motorista(linhas):
        out = []
        for linha in linhas:
            linha = dict(linha)
            nome = linha.pop("_motorista_nome", None)
            if nome is not None:
                mid = mapa_mot.get(nome)
                if not mid:
                    continue                        # motorista desconhecido: pula
                linha["motorista_id"] = mid
            out.append(linha)
        return out

    # 2) veiculos (carretas já vêm antes dos cavalos na lista)
    if contar("veiculos") == 0:
        for v in resolver_motorista(dados["veiculos"]):
            supabase.table("veiculos").insert(v).execute()
        print(f"veiculos: {len(dados['veiculos'])} inseridos")
    else:
        print("veiculos: já tem dados, pulando")

    # 3) demais tabelas, em lotes
    for tabela in ORDEM_INSERCAO[2:]:
        if contar(tabela) > 0:
            print(f"{tabela}: já tem dados, pulando")
            continue
        linhas = resolver_motorista(dados[tabela])
        for i in range(0, len(linhas), 200):
            lote = linhas[i:i + 200]
            if lote:
                supabase.table(tabela).insert(lote).execute()
        print(f"{tabela}: {len(linhas)} inseridos")

    print("\nMigração concluída. Confira no painel do Supabase (Table Editor).")


def main():
    try:
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    except Exception:
        pass

    dry = "--dry-run" in sys.argv
    recarregar = "--recarregar" in sys.argv

    from fontes_planilhas import PASTA_MOTORISTAS, XLSX_MOVIMENTACAO, XLSX_CONTAS_FIXAS
    for arquivo in (XLSX_POSTO, XLSX_FROTA, XLSX_MOVIMENTACAO,
                    XLSX_CONTAS_FIXAS, PASTA_MOTORISTAS):
        if not arquivo.exists():
            sys.exit(f"Arquivo não encontrado: {arquivo}")

    print("Lendo fontes de dados...")
    dados, blocos_sem_dono = preparar_tudo()
    resumo(dados, blocos_sem_dono)

    if dry:
        print("\n--dry-run: nada foi gravado. Rode sem --dry-run para inserir.")
        return
    inserir_tudo(dados, recarregar)


if __name__ == "__main__":
    main()
