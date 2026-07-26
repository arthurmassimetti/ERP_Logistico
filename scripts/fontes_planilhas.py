"""
Le as planilhas originais e monta a MESMA estrutura de dados que o antigo
data.js entregava — que foi apagado quando o sistema passou a viver so no
Supabase (24/07/2026).

Isso e o que devolve reprodutibilidade a migracao: com este modulo, o
migrar.py volta a rodar do zero a partir das planilhas de verdade, sem
depender de nenhum arquivo intermediario.

FONTES
  drive-download-.../<MOTORISTA> 2026.xlsx   -> motoristas (nome/cpf/rg) + fretes
  MOVIMENTAÇÃO DIARIA.xlsx                   -> veiculos (km, media, tacografo,
                                                MCT/ANTT, par cavalo-carreta),
                                                contas a pagar avulsas, saldos
  CONTAS FIXAS.xlsx                          -> contas fixas (pessoal + empresa)
  ROTEIRO DIARIO 2026.xlsx                   -> roteiro do dia
  POSTO 2026.xlsx                            -> vinculo motorista <-> veiculo
                                                (os titulos de bloco sao
                                                 "FULANO - PLACA")

O QUE NAO E RECONSTRUIVEL DAQUI
  vales -> na planilha do motorista existem so como anotacao em texto livre no
           rodape ("12/07 1.000,00 VALE", "20/07 PG 500,00 DO VALE"). Nao da
           para parsear com seguranca, e inventar seria pior que faltar.
           Os vales ja lancados vivem no banco e nos backups (backup_banco.py).
"""
import re
from datetime import date, datetime
from pathlib import Path

import openpyxl

RAIZ = Path(__file__).resolve().parent.parent
FONTES = RAIZ / "fontes"

# A pasta fontes/ e a fonte de verdade (ver fontes/LEIA-ME.md). O caminho antigo
# do Google Drive fica como reserva so para quem ainda nao consolidou.
PASTA_MOTORISTAS = FONTES / "motoristas"
if not PASTA_MOTORISTAS.exists():
    PASTA_MOTORISTAS = RAIZ.parent / "drive-download-20260717T145027Z-1-001"


def _fonte(nome):
    """Prefere fontes/, cai para a raiz do projeto se ainda nao foi consolidado."""
    p = FONTES / nome
    return p if p.exists() else RAIZ / nome


XLSX_MOVIMENTACAO = _fonte("MOVIMENTAÇÃO DIARIA.xlsx")
XLSX_CONTAS_FIXAS = _fonte("CONTAS FIXAS.xlsx")
XLSX_ROTEIRO = _fonte("ROTEIRO DIARIO 2026.xlsx")
XLSX_POSTO = _fonte("POSTO 2026.xlsx")

# arquivo -> nome como o motorista e conhecido no sistema
ARQUIVO_MOTORISTA = {
    "RAIMUNDO 2026.xlsx":        "RAIMUNDO MANOEL",
    "ALESSANDRO VEIGA 2026.xlsx": "ALESSANDRO VEIGA",
    "JOENISON PEREIRA 2026.xlsx": "JOENISON PEREIRA",
    "ARLINDO 2026.xlsx":         "ARLINDO TADEU",
    "EDSON SILVA 2026.xlsx":     "EDSON SILVA",
    "FABIO 2026.xlsx":           "FABIO NOGUEIRA",
    "WILLIAM 2026.xlsx":         "WILLIAM OLIVEIRA",
}

MESES = {"JANEIRO", "FEVEREIRO", "MARÇO", "MARCO", "ABRIL", "MAIO", "JUNHO",
         "JULHO", "AGOSTO", "SETEMBRO", "OUTUBRO", "NOVEMBRO", "DEZEMBRO"}

# colunas da planilha de motorista (conferido nas 7)
C_DATA, C_ORIGEM, C_DESTINO, C_TRANSP = 0, 1, 2, 3
C_FRETE, C_ADIANT, C_DIARIA, C_SALDO, C_COMISS = 4, 5, 6, 7, 8
C_PAGTO, C_BANCO, C_CIOT, C_PEDAGIO, C_PEDAGIO_VIA, C_OBS = 9, 10, 11, 12, 13, 14

# grafias erradas encontradas nas planilhas -> forma canonica
ALIAS_PLACA = {
    "NNX3195": "NNX3I95",   # 1 no lugar do I
    "EYD9D85": "EYV9D85",   # D no lugar do V (bloco de médias da MOVIMENTAÇÃO)
}

RE_PLACA = re.compile(r"^[A-Z]{3}\d[A-Z0-9]\d{2}$")


def eh_placa(v):
    """Evita que cabeçalhos ('CAVALOS', 'CARRETAS', 'PLACA') virem veículo."""
    return bool(v) and bool(RE_PLACA.match(str(v)))

# Regra da frota, informada pelo dono: SO estas placas sao cavalo (caminhao
# trator). Todo o resto e carreta. Isso e necessario porque a propria planilha
# lista RKK4I86 na coluna "CAVALOS" do bloco de pares, e ele e carreta —
# foi essa linha que fez a migracao original classifica-lo errado.
CAVALOS = {"AXN7H33", "EHH9B50", "GAK3400", "NNX3I95", "NNZ5370", "EYV9D85", "MFU2J59"}


def placa_norm(p):
    if not p:
        return None
    p = re.sub(r"\s+", "", str(p)).upper()
    return ALIAS_PLACA.get(p, p)


def d_iso(v):
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, date):
        return v.isoformat()
    return None


def n(v):
    return float(v) if isinstance(v, (int, float)) else None


def t(v):
    if v is None:
        return None
    s = str(v).strip()
    return s or None


# ---------------------------------------------------------------------------
# motoristas + fretes (7 planilhas)
# ---------------------------------------------------------------------------
RE_CPF = re.compile(r"(\d{3}\.\d{3}\.\d{3}-\d{2})")
RE_RG = re.compile(r"RG\s+([\d.\-]+(?:\s+\w+\s+\w+)?)", re.I)


def _doc_do_motorista(wb):
    """Procura CPF e RG nas celulas de qualquer aba (ficam no rodape)."""
    cpf = rg = None
    for aba in wb.sheetnames:
        for row in wb[aba].iter_rows(values_only=True):
            for c in row:
                if c is None:
                    continue
                s = str(c)
                if cpf is None:
                    m = RE_CPF.search(s)
                    if m:
                        cpf = m.group(1)
                if rg is None:
                    m = RE_RG.search(s)
                    if m:
                        achado = m.group(1).strip()
                        # na planilha do WILLIAM o RG e so um traco (nao preenchido)
                        rg = achado if achado.strip("-. ") else None
            if cpf and rg:
                return cpf, rg
    return cpf, rg


def ler_motoristas_e_fretes(mapa_mot_veic=None):
    """Retorna (motoristas, fretes) no formato que o migrar.py espera."""
    mapa_mot_veic = mapa_mot_veic or {}
    motoristas, fretes = [], []

    for arquivo, nome in ARQUIVO_MOTORISTA.items():
        caminho = PASTA_MOTORISTAS / arquivo
        if not caminho.exists():
            raise FileNotFoundError(f"planilha do motorista não encontrada: {caminho}")
        wb = openpyxl.load_workbook(caminho, data_only=True, read_only=True)

        cpf, rg = _doc_do_motorista(wb)
        motoristas.append({
            "nome": nome, "cpf": cpf, "rg": rg,
            "veiculo": mapa_mot_veic.get(nome),
        })

        for aba in wb.sheetnames:
            if aba.strip().upper() not in MESES:
                continue
            for row in wb[aba].iter_rows(min_row=2, values_only=True):
                if not row or len(row) <= C_FRETE:
                    continue
                data = d_iso(row[C_DATA])
                frete = n(row[C_FRETE])
                # linha só é frete se tem data E valor; o resto é cabeçalho,
                # subtotal ou anotação livre do rodapé
                if not data or frete is None:
                    continue
                pega = lambda i: row[i] if len(row) > i else None
                fretes.append({
                    "data": data,
                    "motorista": nome,
                    "veiculo": mapa_mot_veic.get(nome),
                    "origem": t(row[C_ORIGEM]),
                    "destino": t(row[C_DESTINO]),
                    "transportadora": t(row[C_TRANSP]),
                    "frete": frete,
                    "adiant": n(pega(C_ADIANT)),
                    "diaria": n(pega(C_DIARIA)),
                    "saldo": n(pega(C_SALDO)),
                    "comissao": n(pega(C_COMISS)),
                    "pagto": pega(C_PAGTO),
                    "banco": t(pega(C_BANCO)),
                    "ciot": t(pega(C_CIOT)),
                    "pedagio": n(pega(C_PEDAGIO)),
                    "pedagioVia": t(pega(C_PEDAGIO_VIA)),
                    "obs": t(pega(C_OBS)),
                })
        wb.close()

    return motoristas, fretes


# ---------------------------------------------------------------------------
# vinculo motorista <-> veiculo (titulos de bloco do POSTO: "FULANO - PLACA")
# ---------------------------------------------------------------------------
def ler_mapa_motorista_veiculo():
    """Titulos de bloco do POSTO no formato "FULANO - PLACA".

    A planilha e irregular: grafa "WILLIAN" no lugar de WILLIAM e abrevia a
    placa em "JOENISON - NNX". Por isso o nome casa pelos 5 primeiros
    caracteres e a placa aceita prefixo, desde que resolva para um cavalo
    conhecido sem ambiguidade. Motoristas antigos que aparecem nos titulos
    (HEVERTON, SERGIO, PAI THOMAZ) simplesmente nao casam com ninguem.
    """
    mapa = {}
    if not XLSX_POSTO.exists():
        return mapa
    wb = openpyxl.load_workbook(XLSX_POSTO, data_only=True, read_only=True)
    nomes = list(ARQUIVO_MOTORISTA.values())
    for aba in wb.sheetnames:
        for row in wb[aba].iter_rows(values_only=True):
            titulo = t(row[0]) if row else None
            if not titulo or "-" not in titulo:
                continue
            alvo = titulo.upper()
            m = re.search(r"[A-Z]{3}\s?\d[A-Z0-9]\d{2}", alvo)
            if m:
                placa = placa_norm(m.group(0))
            else:
                # placa abreviada: resolve por prefixo, so se for unica
                sufixo = placa_norm(alvo.split("-", 1)[1]) or ""
                candidatas = [c for c in CAVALOS if sufixo and c.startswith(sufixo)]
                if len(candidatas) != 1:
                    continue
                placa = candidatas[0]
            if placa not in CAVALOS:
                continue
            for nome in nomes:
                if nome.split()[0][:5] in alvo:
                    mapa.setdefault(nome, placa)
    wb.close()
    return mapa


# ---------------------------------------------------------------------------
# MOVIMENTAÇÃO DIARIA — veiculos, contas a pagar, saldos
# ---------------------------------------------------------------------------
def _linhas_movimentacao():
    wb = openpyxl.load_workbook(XLSX_MOVIMENTACAO, data_only=True, read_only=True)
    linhas = list(wb["2026"].iter_rows(values_only=True))
    wb.close()
    return linhas


def ler_veiculos(mapa_mot_veic):
    """Cavalos e carretas a partir dos varios blocos da MOVIMENTAÇÃO DIARIA."""
    linhas = _linhas_movimentacao()
    col = lambda row, i: row[i] if len(row) > i else None

    km, media, taco, mct, antt_cav, antt_car, pares = {}, {}, {}, {}, {}, {}, []

    for row in linhas:
        if not row:
            continue
        # bloco 1: PLACA | . | ATUAL | . | TROCA
        p = placa_norm(col(row, 0))
        if p and n(col(row, 2)):
            km[p] = {"kmAtual": int(n(col(row, 2))), "kmTroca": int(n(col(row, 4)) or 0) or None}
        # bloco 2: PLACA | . | MEDIA | DATA
        p = placa_norm(col(row, 8))
        if p and n(col(row, 10)):
            media[p] = {"media": n(col(row, 10)), "mediaData": d_iso(col(row, 11))}
        # bloco 3: PLACA | DATA VENCIMENTO | obs
        p = placa_norm(col(row, 32))
        if p and d_iso(col(row, 34)):
            taco[p] = {"tacografo": d_iso(col(row, 34)), "tacografoObs": t(col(row, 35))}
        # bloco 4: PLACA | MCT | status
        p = placa_norm(col(row, 40))
        if p and t(col(row, 41)):
            mct[p] = {"mct": t(col(row, 41)), "mctStatus": t(col(row, 42))}
        # bloco 5/6: ANTT de cavalo e de carreta
        p = placa_norm(col(row, 43))
        if eh_placa(p) and t(col(row, 44)):
            antt_cav[p] = {"antt": t(col(row, 44)), "anttNum": t(col(row, 45))}
        p = placa_norm(col(row, 47))
        if eh_placa(p) and t(col(row, 48)):
            antt_car[p] = {"antt": t(col(row, 48)), "anttNum": t(col(row, 49))}
        # bloco 7: par CAVALOS | CARRETAS. A coluna "CAVALOS" da planilha nao
        # e confiavel (lista RKK4I86, que e carreta) — vale a regra em CAVALOS.
        cav, car = placa_norm(col(row, 58)), placa_norm(col(row, 60))
        cav = cav if eh_placa(cav) else None
        car = car if eh_placa(car) else None
        if cav and cav in CAVALOS:
            pares.append((cav, car))
        elif cav:
            # placa listada na coluna "CAVALOS" que na verdade e carreta
            pares.append((None, cav))
        if car and not cav:
            pares.append((None, car))

    veiculo_do_motorista = {v: k for k, v in mapa_mot_veic.items()}
    cavalos, carretas = [], []
    vistos_carreta = set()

    for cav, car in pares:
        if cav is None:
            if car and car not in vistos_carreta:
                vistos_carreta.add(car)
                carretas.append(car)
            continue
        d = {"placa": cav, "carreta": car, "motorista": veiculo_do_motorista.get(cav)}
        d.update(km.get(cav, {}))
        d.update(media.get(cav, {}))
        d.update(taco.get(cav, {}))
        d.update(mct.get(cav, {}))
        d.update(antt_cav.get(cav, {}))
        cavalos.append(d)
        if car and car not in vistos_carreta:
            vistos_carreta.add(car)
            carretas.append(car)

    # carretas que aparecem so no bloco de ANTT de carreta
    for p in antt_car:
        if p not in vistos_carreta:
            vistos_carreta.add(p)
            carretas.append(p)

    return carretas, cavalos


def ler_contas_pagar():
    """Bloco financeiro (colunas 21-24) + contas futuras (25-28)."""
    from importar_avulsas import ler_planilha  # mesma logica, ja validada
    return ler_planilha()


def ler_saldos():
    """Linha 'SALDO:' do bloco de saldos: rotulos na linha de cima, valores embaixo."""
    linhas = _linhas_movimentacao()
    col = lambda row, i: row[i] if len(row) > i else None
    rotulos, saldos = [], []
    for i, row in enumerate(linhas):
        if not row:
            continue
        if any(t(c) and str(c).strip().upper() == "SALDO:" for c in row):
            anterior = linhas[i - 1] if i else []
            for j in (16, 17, 18, 19):
                nome = t(col(anterior, j))
                valor = n(col(row, j))
                if nome and valor is not None:
                    rotulos.append(nome)
                    saldos.append({"banco": nome, "saldo": valor})
            break
    return saldos


# ---------------------------------------------------------------------------
# CONTAS FIXAS
# ---------------------------------------------------------------------------
def ler_contas_fixas():
    wb = openpyxl.load_workbook(XLSX_CONTAS_FIXAS, data_only=True, read_only=True)
    out = []
    for aba in wb.sheetnames:
        origem = "pessoal" if aba.strip().upper() == "PESSOAL" else "empresa"
        for row in wb[aba].iter_rows(min_row=3, values_only=True):
            if not row or len(row) < 4:
                continue
            dia, desc, valor = row[1], t(row[2]), n(row[3])
            if not desc or valor is None or not isinstance(dia, (int, float)):
                continue
            # coluna 4 e "MENSAL" ou uma data de fim de contrato
            quarta = row[4] if len(row) > 4 else None
            fim = d_iso(quarta)
            out.append({
                "dia": int(dia),
                "descricao": desc,
                "valor": valor,
                "recorrencia": None if fim else (t(quarta) or "MENSAL"),
                "origem": origem,
                "forma": t(row[5]) if len(row) > 5 else None,
                "fim": fim,
                "ativa": True,
                "pendente": "FOLHA" in desc.upper(),
            })
    wb.close()
    return out


# ---------------------------------------------------------------------------
# ROTEIRO
# ---------------------------------------------------------------------------
def ler_roteiro():
    wb = openpyxl.load_workbook(XLSX_ROTEIRO, data_only=True, read_only=True)
    out = []
    nomes = list(ARQUIVO_MOTORISTA.values())
    # colunas (conferido): [1]ordem [2]onde está [3]motorista [4]origem UF
    #                      [5]destino UF [6]rota [7]transportadora [8]status
    for row in wb["ROTEIRO"].iter_rows(values_only=True):
        if not row or len(row) < 4:
            continue
        ordem = row[1] if isinstance(row[1], (int, float)) else None
        nome_cel = t(row[3])
        if ordem is None or not nome_cel:
            continue
        nome = next((x for x in nomes if x.split()[0][:5] in nome_cel.upper()), nome_cel)
        pega = lambda i: t(row[i]) if len(row) > i else None
        origem_uf, destino_uf = pega(4), pega(5)
        out.append({
            "ordem": int(ordem),
            "motorista": nome,
            "posicao": f"{origem_uf} → {destino_uf}" if origem_uf and destino_uf else (origem_uf or ""),
            "rota": pega(6),
            "status": pega(8),
        })
    wb.close()
    return out


# ---------------------------------------------------------------------------
# entrada principal — devolve o mesmo dicionario que o data.js tinha
# ---------------------------------------------------------------------------
def carregar_planilhas():
    mapa = ler_mapa_motorista_veiculo()
    motoristas, fretes = ler_motoristas_e_fretes(mapa)
    carretas, cavalos = ler_veiculos(mapa)
    return {
        "hoje": date.today().isoformat(),
        "motoristas": motoristas,
        "carretas": carretas,
        "veiculos": cavalos,
        "fretes": fretes,
        "roteiro": ler_roteiro(),
        "contasFixas": ler_contas_fixas(),
        "financeiro": {
            "contasPagar": ler_contas_pagar(),
            "saldos": ler_saldos(),
        },
        # ver docstring do modulo: vales so existem como texto livre na planilha
        "vales": [],
    }


if __name__ == "__main__":
    import sys
    sys.stdout.reconfigure(encoding="utf-8")
    db = carregar_planilhas()
    print("Lido das planilhas:")
    print(f"  motoristas   : {len(db['motoristas'])}")
    print(f"  cavalos      : {len(db['veiculos'])}")
    print(f"  carretas     : {len(db['carretas'])}")
    print(f"  fretes       : {len(db['fretes'])}  "
          f"(R$ {sum(f['frete'] for f in db['fretes']):,.2f})")
    print(f"  roteiro      : {len(db['roteiro'])}")
    print(f"  contas fixas : {len(db['contasFixas'])}")
    print(f"  contas pagar : {len(db['financeiro']['contasPagar'])}")
    print(f"  saldos       : {len(db['financeiro']['saldos'])}  {db['financeiro']['saldos']}")
    print(f"  vales        : {len(db['vales'])}  (não reconstruível — ver docstring)")
    for m in db["motoristas"]:
        print(f"    {m['nome']:20} cpf={m['cpf']} rg={str(m['rg'])[:14]:14} veiculo={m['veiculo']}")
